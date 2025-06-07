import numpy as np
import pandas as pd
import fitz  # PyMuPDF
from langchain.text_splitter import RecursiveCharacterTextSplitter
from transformers import pipeline as hf_pipeline, AutoTokenizer
from sentence_transformers import SentenceTransformer
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
import time
import os
import torch
from concurrent.futures import ThreadPoolExecutor
from sklearn.metrics.pairwise import cosine_similarity
from collections import OrderedDict
import warnings
from docx import Document
import math

os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
warnings.filterwarnings("ignore", category=FutureWarning)
torch.set_num_threads(4)

# Configuration - Optimized for 800-page book
QA_MODEL = "deepset/roberta-base-squad2"  # More robust QA model
EMBEDDING_MODEL = "all-mpnet-base-v2"    # Better semantic embeddings
SUMMARY_MODEL = "facebook/bart-large-cnn" # Higher quality summaries
CHUNK_SIZE = 500
CHUNK_OVERLAP = 100
MIN_CONFIDENCE = 0.2
SIMILARITY_THRESHOLD = 0.4
MAX_DOC_PAGES = 1000
BATCH_SIZE = 16
SUMMARY_MAX_LENGTH = 600
SUMMARY_MIN_LENGTH = 200
SUMMARY_TARGET_RATIO = 0.85

def extract_text_optimized(pdf_paths):
    """Optimized text extraction with parallel processing"""
    def process_page(page, pdf_path):
        text = page.get_text("text")
        return {
            "text": re.sub(r'\s+', ' ', text).strip(),
            "page": page.number + 1,
            "source": os.path.basename(pdf_path)
        }

    text_data = []
    for pdf_path in (pdf_paths if isinstance(pdf_paths, list) else [pdf_paths]):
        try:
            with fitz.open(pdf_path) as doc:
                with ThreadPoolExecutor() as executor:
                    pages = list(executor.map(
                        lambda i: process_page(doc[i], pdf_path),
                        range(min(MAX_DOC_PAGES, len(doc)))))
                    text_data.extend([p for p in pages if p["text"]])
        except Exception as e:
            print(f"Error processing {pdf_path}: {e}")
    return text_data

class QAPipeline:
    def __init__(self):
        self.device = 0 if torch.cuda.is_available() else -1
        self.model = hf_pipeline(
            "question-answering",
            model=QA_MODEL,
            device=self.device
        )

    def get_answers(self, question, contexts):
        inputs = [{"question": question, "context": c} for c in contexts]
        return self.model(inputs, batch_size=BATCH_SIZE)

class SummaryGenerator:
    def __init__(self):
        self.device = 0 if torch.cuda.is_available() else -1
        self.model = hf_pipeline(
            "summarization",
            model=SUMMARY_MODEL,
            device=self.device
        )
        self.tokenizer = AutoTokenizer.from_pretrained(SUMMARY_MODEL)

    def summarize(self, text, max_length=None, min_length=None):
        try:
            if not text.strip():
                return ""
            text = re.sub(r'\s+', ' ', text.replace('\n', ' ')).strip()
            sentences = [s.strip() for s in re.split(r'(?<=[.!?])\s+', text) if s.strip()]
            if not sentences:
                return ""
            clean_text = ' '.join(sentences[:30])

            if min_length is None:
                min_length = SUMMARY_MIN_LENGTH
            if max_length is None:
                input_tokens = len(self.tokenizer.tokenize(clean_text))
                max_length = min(
                    int(input_tokens * SUMMARY_TARGET_RATIO),
                    SUMMARY_MAX_LENGTH
                )
                max_length = max(max_length, min_length + 50)

            summary = self.model(
                clean_text,
                max_length=max_length,
                min_length=min_length,
                truncation=True,
                num_beams=4,
                length_penalty=1.0,
                no_repeat_ngram_size=3
            )[0]['summary_text']

            if not summary.endswith(('.','!','?')):
                summary = summary.rsplit('. ', 1)[0] + '.' if '. ' in summary else summary + '.'
            return summary

        except Exception as e:
            print(f"Summarization error: {e}")
            return text[:400] + ('...' if len(text) > 400 else '')

class MaturityAnalyzer:
    def __init__(self):
        self.embedder = SentenceTransformer(EMBEDDING_MODEL)
        self.summarizer = SummaryGenerator()

    def load_maturity_framework(self, docx_path):
        doc = Document(docx_path)
        options_map = {}
        desc_map = {}
        for table in doc.tables:
            if len(table.rows) < 2:
                continue
            headers = [self.clean_text(cell.text) for cell in table.rows[0].cells]
            for row in table.rows[1:]:
                cells = [self.clean_text(cell.text) for cell in row.cells]
                if len(cells) < 2:
                    continue
                question = cells[0]
                if not question:
                    continue
                is_question = any(
                    question.lower().startswith(q_word)
                    for q_word in ['what', 'how', 'does', 'is', 'to what']
                )
                if not is_question:
                    continue
                options = []
                descriptions = []
                for i in range(1, min(len(cells), len(headers))):
                    if i < len(headers) and cells[i]:
                        level = headers[i].split(":")[0].strip()
                        full_desc = f"{headers[i]}: {cells[i]}"
                        options.append(level)
                        descriptions.append(full_desc)
                if question not in options_map:
                    options_map[question] = options
                    desc_map[question] = descriptions
        return list(options_map.keys()), options_map, desc_map

    def clean_text(self, text):
        text = re.sub(r'[\x00-\x1F\x7F-\x9F]', ' ', str(text))
        text = text.replace('‑', '-').replace('–', '-')
        return text.strip()

    def match_maturity(self, question, summary, options_map, desc_map):
        try:
            # Enhanced question matching
            matched_question = question
            if question not in options_map:
                normalized_question = re.sub(r'[^\w\s]', '', question.lower()).strip()
                matched_question = next(
                    (q for q in options_map.keys() 
                    if re.sub(r'[^\w\s]', '', q.lower()).strip() == normalized_question),
                    None
                )
            
            if not matched_question:
                return {
                    'level': 'N/A',
                    'description': 'Question not in framework',
                    'confidence': 0
                }

            # Summarize options for better matching
            summarized_options = [
                self.summarizer.summarize(desc, max_length=150) 
                for desc in desc_map[matched_question]
            ]
            texts = [summary] + summarized_options
            embeddings = self.embedder.encode(texts)
            
            summary_embedding = embeddings[0]
            option_embeddings = embeddings[1:]
            similarities = cosine_similarity([summary_embedding], option_embeddings)[0]
            
            best_idx = similarities.argmax()
            best_similarity = similarities[best_idx]
            
            # Boost confidence for higher levels when appropriate
            if best_similarity > 0.6 and int(options_map[matched_question][best_idx]) >= 3:
                best_similarity = min(best_similarity * 1.2, 1.0)
            
            return {
                'level': options_map[matched_question][best_idx],
                'description': self.clean_text(desc_map[matched_question][best_idx]),
                'confidence': float(best_similarity)
            }

        except Exception as e:
            print(f"Maturity matching error: {e}")
            return {
                'level': 'Error',
                'description': 'Analysis failed',
                'confidence': 0
            }

def save_final_excel(results, output_path, sample_chunks=None, all_questions=None, maturity_docx_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    
    summarizer = SummaryGenerator()
    analyzer = MaturityAnalyzer()
    
    options_map, desc_map = {}, {}
    if maturity_docx_path:
        try:
            _, options_map, desc_map = analyzer.load_maturity_framework(maturity_docx_path)
        except Exception as e:
            print(f"Failed to load maturity framework: {e}")
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    wrap_align = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    question_groups = OrderedDict()
    for result in results:
        question = result['question']
        if question not in question_groups:
            question_groups[question] = []
        question_groups[question].append(result)
    
    ws.append(["Question"] + list(question_groups.keys()))
    
    for col, question in enumerate(question_groups.keys(), start=2):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(question) * 0.7, 20), 50)
    
    ws.cell(row=2, column=1, value="Maturity Description")
    for col_idx, (question, group_results) in enumerate(question_groups.items(), start=2):
        contexts = [r['context'] for r in group_results]
        summary = summarizer.summarize("\n".join(contexts))
        maturity = analyzer.match_maturity(question, summary, options_map, desc_map)
        ws.cell(row=2, column=col_idx, value=maturity['description'])
    
    ws.cell(row=3, column=1, value="Summary")
    for col_idx, (question, group_results) in enumerate(question_groups.items(), start=2):
        contexts = [r['context'] for r in group_results]
        summary = summarizer.summarize("\n".join(contexts))
        ws.cell(row=3, column=col_idx, value=summary)
    
    max_chunks = max(len(group) for group in question_groups.values())
    for chunk_num in range(max_chunks):
        row_num = 4 + chunk_num
        ws.cell(row=row_num, column=1, value=f"Extract Statement {chunk_num+1}")
        for col_idx, (question, group_results) in enumerate(question_groups.items(), start=2):
            if chunk_num < len(group_results):
                result = group_results[chunk_num]
                clean_context = analyzer.clean_text(result['context'])
                chunk_text = (
                    f"[Page {result['page']} | {result['source']}]:\n"
                    f"{clean_context}"
                )
                ws.cell(row=row_num, column=col_idx, value=chunk_text)
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
    
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                col_width = ws.column_dimensions[cell.column_letter].width
                if col_width > 0:
                    approx_chars_per_line = max(5, col_width * 1.8)
                    lines = max(lines, math.ceil(len(str(cell.value)) / approx_chars_per_line))
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max(20, max_lines * 15)
    
    ws.freeze_panes = 'B2'
    wb.save(output_path)

def process_documents(pdf_paths, questions, output_path, maturity_docx_path=None):
    print("Starting optimized QA processing...")
    start_time = time.time()
    
    maturity_analyzer = MaturityAnalyzer() if maturity_docx_path else None
    unique_questions = list(OrderedDict.fromkeys(questions.keys()))
    if len(unique_questions) != len(questions):
        print(f"Duplicate questions detected. Using {len(unique_questions)} unique questions.")
        questions = OrderedDict((q, q) for q in unique_questions)
    
    print("Loading models...")
    embed_model = SentenceTransformer(EMBEDDING_MODEL)
    qa_pipeline = QAPipeline()
    
    print("Extracting text...")
    text_data = extract_text_optimized(pdf_paths)
    print(f"Found {len(text_data)} pages")
    
    print("Splitting text...")
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", "! ", "? ", "; ", " "]
    )
    chunks = []
    for data in text_data:
        chunks.extend([{"text": chunk, "page": data["page"], "source": data["source"]}
                      for chunk in splitter.split_text(data["text"])])
    print(f"Created {len(chunks)} chunks")
    
    print("Generating embeddings...")
    embeddings = embed_model.encode(
        [chunk["text"] for chunk in chunks],
        batch_size=BATCH_SIZE,
        show_progress_bar=True
    )
    
    print("Building index...")
    embeddings = np.array(embeddings).astype('float32')
    faiss.normalize_L2(embeddings)
    index = faiss.IndexFlatIP(embeddings.shape[1])
    index.add(embeddings)
    
    print(f"\nProcessing {len(questions)} questions...")
    results = []
    question_embeds = embed_model.encode(list(questions.keys()))
    
    for q_idx, (question, q_embed) in enumerate(zip(questions.keys(), question_embeds)):
        print(f"\nProcessing question {q_idx+1}/{len(questions)}: {question[:50]}...")
        q_embed = q_embed.reshape(1, -1).astype('float32')
        faiss.normalize_L2(q_embed)
        distances, indices = index.search(q_embed, len(chunks))
        
        relevant_chunks = [
            chunks[i] for i in indices[0] 
            if (1 - distances[0][i]) >= SIMILARITY_THRESHOLD
        ][:30]
        
        print(f"Found {len(relevant_chunks)} relevant chunks")
        
        if not relevant_chunks:
            results.append({
                'question': question,
                'answer': "NO RELEVANT CONTEXT FOUND",
                'score': 0,
                'page': 0,
                'source': "",
                'context': ""
            })
            continue
            
        contexts = [c["text"] for c in relevant_chunks]
        answers = qa_pipeline.get_answers(question, contexts)
        
        found_answers = False
        for ans, chunk in zip(answers, relevant_chunks):
            if ans['score'] >= MIN_CONFIDENCE:
                found_answers = True
                results.append({
                    'question': question,
                    'answer': ans['answer'],
                    'score': ans['score'],
                    'page': chunk['page'],
                    'source': chunk['source'],
                    'context': chunk['text']
                })
        
        if not found_answers:
            results.append({
                'question': question,
                'answer': "NO CONFIDENT ANSWER FOUND",
                'score': 0,
                'page': relevant_chunks[0]['page'] if relevant_chunks else 0,
                'source': relevant_chunks[0]['source'] if relevant_chunks else "",
                'context': relevant_chunks[0]['text'] if relevant_chunks else ""
            })
    
    processed_questions = {r['question'] for r in results}
    missing_questions = set(questions.keys()) - processed_questions
    if missing_questions:
        print(f"\nMissing results for {len(missing_questions)} questions:")
        for q in missing_questions:
            results.append({
                'question': q,
                'answer': "PROCESSING ERROR",
                'score': 0,
                'page': 0,
                'source': "",
                'context': ""
            })
    
    print("Saving results...")
    save_final_excel(
        results,
        output_path,
        maturity_docx_path=maturity_docx_path,
        sample_chunks=[chunk['text'] for chunk in chunks[:10]],
        all_questions=list(questions.keys())
    )
    
    print(f"\nCompleted in {time.time()-start_time:.1f} seconds")

if __name__ == "__main__":
    pdf_paths = "C:/Users/gauri/Documents/Git_Clones/python_codes/pdf extraction/book.pdf"  # Replace with your 800-page book PDF
    output_path = "book_analysis.xlsx"
    maturity_docx_path = "maturity_framework.docx"  # Replace with your maturity framework
    
    questions_list = [
        "What is the company’s purpose/vision/mission?",
        "To what extent use and how does the company manage external contractors / non-permanent /temporary employees?",
        "How does the company interact with its different stakeholders (customers, users, suppliers, employees, regulators, investors, government, society)?",
        "Is the company leveraging different disruptive and emerging technologies",
        "What are the key Metrics / KPIs / key performance indicators being tracked related to the innovation portfolio and overall business performance?",
        "Does the organization tolerate failure and encourage risk-taking?"
    ]
    
    questions_dict = OrderedDict((q, q) for q in questions_list)
    process_documents(
        pdf_paths=pdf_paths,
        questions=questions_dict,
        output_path=output_path,
        maturity_docx_path=maturity_docx_path
    )